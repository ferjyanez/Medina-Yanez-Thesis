import tensorflow as tf
import openpyxl as oxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import scipy.stats as ss


from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter


import os
duration = 1  # second
freq = 440  # Hz

#initial variables 
l_rate_0 = 0.001
epochs = 10000

opcion = 6

wb = oxl.Workbook()
if(opcion == 1):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/MyData.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:200]
  
  df_X = df.drop(df.columns[[4, 5]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 3]], axis=1)

elif(opcion == 2):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/FB-HP.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:99]

  df_X = df.drop(df.columns[[3, 4]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

elif(opcion == 3):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/hate_crimes.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:50]

  df_X = df.drop(df.columns[[9, 10]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 3, 4, 5, 6, 7, 8]], axis=1)

elif(opcion == 4):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/FB-HP.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:99]

  df_X = df.drop(df.columns[[2, 3, 4]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 4]], axis=1)

elif(opcion == 5):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/Ejemplito.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[2]], axis=1)
  df_Y = df.drop(df.columns[[0, 1]], axis=1)

elif(opcion == 6):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/Ejemplito2.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[2, 3]], axis=1)
  df_Y = df.drop(df.columns[[0, 1]], axis=1)

elif(opcion == 7):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/ejemplito4_.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[3]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

elif(opcion == 8):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/ejemplito5_.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[4]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 3]], axis=1)


X_matrix = np.asmatrix(df_X)
Y_matrix = np.asmatrix(df_Y)
#Sizes
x_var = X_matrix.shape[1] #x_var
y_var = Y_matrix.shape[1] #y_var
n_ind = X_matrix.shape[0] #n_ind

X = tf.placeholder(tf.float32, [None,x_var]) #Inserts a placeholder for a tensor that will be always fed.
Y = tf.placeholder(tf.float32, [None,y_var]) #Inserts a placeholder for a tensor that will be always fed.

W_1 = tf.Variable(tf.truncated_normal([x_var,1], stddev=0.01)) #gets a tensor with the values
W = tf.Variable(tf.truncated_normal([x_var,y_var], stddev=0.01)) #gets a tensor with the values

model = tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W_1),tf.transpose(W_1)),tf.transpose(X)),Y) #Y estimada

#Cost function
cost0 = tf.square(Y - model) #subtracts element-wise the vector and the elevates them to the 2nd power
cost1 = tf.reduce_sum(cost0)

cost2 = 0



def f_W(X1, X2):
  X1[:,0].assign(X2)
  return X1


#cost2 = tf.trace(tf.matmul(tf.transpose(tf.subtract(tf.matmul(tf.transpose(f_W(W, W_1)),f_W(W, W_1)),tf.eye(y_var))),tf.subtract(tf.matmul(tf.transpose(f_W(W, W_1)),f_W(W, W_1)),tf.eye(y_var))))

                         

loss = tf.add(cost1, cost2)

#Optimizer
#global_step = tf.Variable(0, trainable=False)
#l_rate = tf.train.exponential_decay(l_rate_0, global_step, 100000, 0.96, staircase=True)
optimizer = tf.train.GradientDescentOptimizer(l_rate_0)
train = optimizer.minimize(loss)

errors_e = []

#training
#initialise the global variables in order to run the session
init = tf.global_variables_initializer()
sess = tf.Session()
sess.run(init)
for i in range(epochs):
  _, loss_e = sess.run([train, loss], {X: X_matrix, Y: Y_matrix})
  errors_e.append(loss_e/n_ind)


#Get results

final_W, loss_eu = sess.run([W_1, loss], {X:X_matrix, Y:Y_matrix})
print("W:")
print(final_W)
print("loss: %s"%(loss_eu/n_ind))
print(np.trace(final_W.transpose().dot(X_matrix.transpose()).dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W)))
print(final_W.transpose().dot(final_W))

sess.close()

os.system('play --no-show-progress --null --channels 1 synth %s sine %f' % (duration, freq))


print("")
W_AR = [[-0.9209633,0.3896493],[-0.3896493,-0.9209633]]
W_AR = np.asmatrix(W_AR)
final_W = np.asmatrix(final_W)
#print(final_W[2,0]/W_AR[2,0])
print("")
print("")
print(W_AR)
print("")
print(final_W)
print("")
#print("")
#print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(W_AR[:,0])/W_AR[:,0])
#print("")
print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W[:,0])/final_W[:,0])

#plt.subplot(1,1,1)
plt.plot(np.asarray(errors_e))
plt.ylabel("Loss")
plt.show()
#plt.savefig("errors.png")

"""

W_AR = [[0.5544050,0.821733],[-0.1993123,1.2850765],[-0.8080283,0.4934761]]
W_AR = np.asmatrix(W_AR)
print("")
print(W_AR[0:3,0])
print("")
print(np.cross(W_AR[:,0].transpose(), final_W[:,0].transpose()))
print("")
print(W_AR[0,0]/final_W[0,0])
print(W_AR[1,0]/final_W[1,0])
print(W_AR[2,0]/final_W[2,0])




def f(w1, w2, X_matrix, Y_matrix):

  Z = np.zeros((20,20))
  for i in range(20):
    for j in range(20):
      Z[i,j] = (Y_matrix - (np.multiply(np.square(w1[i,j]),X_matrix[:,0])
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,1])).dot(
                  X_matrix[:,0].transpose()
                ).dot(
                  Y_matrix)
              + (np.multiply(np.square(w2[i,j]),X_matrix[:,1]) 
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,0])).dot(
                  X_matrix[:,1].transpose()
                ).dot(
                    Y_matrix)).transpose().dot(
      Y_matrix - (np.multiply(np.square(w1[i,j]),X_matrix[:,0])
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,1])).dot(
                  X_matrix[:,0].transpose()
                ).dot(
                  Y_matrix[:,0])
              + (np.multiply(np.square(w2[i,j]),X_matrix[:,1]) 
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,0])).dot(
                  X_matrix[:,1].transpose()
                ).dot(
                    Y_matrix)
              )
  return Z

# Make data.
fig = plt.figure()
ax = fig.gca(projection='3d')

w1 = np.linspace(-1.3, 1.3, 20)
w2 = np.linspace(-1.3, 1.3, 20)
w1, w2 = np.meshgrid(w1, w2)

Z = np.multiply(f(w1, w2, X_matrix, Y_matrix),(1/n_ind))


# Plot the surface.
surf = ax.plot_surface(w1, w2, Z, cmap = cm.coolwarm,linewidth=0, antialiased=False)

# Customize the z axis.
ax.set_zlim(0, 2.5)
ax.zaxis.set_major_locator(LinearLocator(10))
ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))

# Add a color bar which maps values to colors.
fig.colorbar(surf, shrink=0.5, aspect=5)

plt.show()

 """