# -*- coding: utf-8 -*-
"""
Created on Wed Apr 18 21:48:41 2018

@author: Fernando J. Yanez

Name: RNA-TesisMedinaYanez
Last modification: 11May2018

"""

#import tensorflow as tf
import openpyxl as oxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import scipy.stats as ss

import os
duration = 1  # second
freq = 440  # Hz

#==============================================================================
#Global functions
#==============================================================================
def vec_by_col(X_):
    return np.asmatrix(np.reshape(X_, X_.shape[0]*X_.shape[1], order='F')).transpose()

def per_matrix(m_, n_):
    M = np.zeros((m_*n_,m_*n_))
    for i in range(n_):
        for j in range(m_):
            M[(i*m_)+j][(j*n_)+i] = 1
    return M
#==============================================================================


#==============================================================================
#Global variables
#==============================================================================
#Coeficiente de aprendizaje
#l_rate = 0.02
l_rate_0 = 0.02
dacay_rate = 1

#cal
calibration_epsilon = 1.0
calibration_etha = 5.0
calibration = 1.90


#Numero de corridas maximo para el entrenamiento
epochs = 500
#==============================================================================


#==============================================================================
#Toma de la data
#==============================================================================
##Importa la data de un archivo .xlsx 
wb = oxl.Workbook()
wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/MyData.xlsx')
#wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/datosfer.xlsx')
#wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/ENB2012_data.xlsx')
ws = wb.active
df = pd.DataFrame(ws.values)
#==============================================================================
    

##==============================================================================
##Separacion de la data
##==============================================================================
df = df.drop(df.columns[[0]], axis=1)
df = df[:201]
df = df.drop(df.index[0])

df_X = df.drop(df.columns[[4, 5]], axis=1)
df_Y = df.drop(df.columns[[0, 1, 2, 3]], axis=1)

X = np.asmatrix(df_X, dtype = np.float32)
Y = np.asmatrix(df_Y, dtype = np.float32)

#X = ss.zscore(X, axis=0)                      ####revisar si es 0 o 1
#Y = ss.zscore(Y, axis=0)

""" df = df.drop(df.columns[[0]], axis=1)
df = df[:51]
df = df.drop(df.index[0])

df_X = df.drop(df.columns[[3, 4]], axis=1)
df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

X = np.asmatrix(df_X)
Y = np.asmatrix(df_Y) """

""" df = pd.DataFrame(ws.values)
df = df.drop(df.columns[[10, 11]], axis=1)
df = df[:769]
df = df.drop(df.index[0])

df_X = df.drop(df.columns[[8, 9]], axis=1)
df_Y = df.drop(df.columns[[0, 1, 2, 3, 4, 5, 6, 7]], axis=1)

X = np.asmatrix(df_X)
Y = np.asmatrix(df_Y)
print(X)
print(Y) """
##============================================================================== 


##==============================================================================
##Constants
##==============================================================================
#Sizes
p = X.shape[1] #x_var
q = Y.shape[1] #y_var
n = X.shape[0] #n_ind

#Identity matrices
Ip = np.eye(p)
Iq = np.eye(q)
#==============================================================================


##==============================================================================
##Usefull varibles for learning
##==============================================================================
M1 = np.kron(Y.transpose().dot(X), np.eye(p))
M2 = vec_by_col(Ip).dot(vec_by_col(Iq).transpose())
M3 = per_matrix(p,q)
M4 = np.kron(vec_by_col(Ip).transpose(), Ip)
M5 = np.kron(per_matrix(p,p),Ip).transpose()
M6 = np.kron(per_matrix(p,p),Iq)
M7 = np.kron(vec_by_col(Ip), Iq)
M8 = np.kron(per_matrix(p,q),Iq)
M9 = np.kron(Ip, per_matrix(p,q))
##============================================================================== 


#==============================================================================    
#Red Neuronal Artificial
#==============================================================================
#Creacion de la matriz W de pesos entre la capa de entrada y la capa oculta
W = np.random.normal(0, 0.1, p*q).reshape(p, q)
#W = np.zeros(p*q).reshape(p, q)

#W = np.asmatrix(W)    


aux = True



errors = []
errors2 = []
contError = 1
#Starts the process
for i in range (1,epochs):#numero de corridas
    l_rate = l_rate_0/((1+dacay_rate)*i)
    #Otra forma
    #l_rate = l_rate_0*((0.95)**i)
    c_errors = 0
    if(contError==0):#condicion de parada         
        break
    else:
        contError = 0
        c_errors = 0        
        for j in range(0,n):#number of entrances of data
            #Calculates
            Ycalc = X[j].dot(W.dot(W.transpose().dot(X.transpose().dot(Y))))

            e = (Ycalc - Y[j]).transpose()
            if(np.math.isnan(np.sum(e))):
                print("es nan")
                break

            #aux_e = np.abs((Ycalc/Y[j])-1)
            #aux_errors = np.sum(aux_e/q)
            aux_e = e.dot(e.transpose())
            aux_errors = np.sum(aux_e)
            c_errors = c_errors + aux_errors

    #==============================================================================


    #==============================================================================       
            #Learning process
    #==============================================================================
            #Tolerance for the error of 5% over the non-normalized data
            #######revisar si np.sqrt(aux_errors)/np.sum(Y[j]) es una buena forma de medir el error porcentual
            if(False):
            #if(np.sqrt(aux_errors)/np.sum(Y[j])>0.975 and np.sqrt(aux_errors)/np.sum(Y[j])<1.025):
                #If the predicted value is within the 5% error, don't learn                        
                delta1 = np.zeros(p*q).reshape((p,q))
                delta2 = np.zeros(p*q).reshape((p,q))
                delta3 = np.zeros(p*q).reshape((p,q))
                if(aux):
                    aux = False
                    print("entro al cero error:")
                    print(delta1)
                    print("epoch")
                    print(i)
                    print("ind")
                    print(j)
            else:
                contError = contError + 1

                delta1 = -np.kron(2*e,Ip).transpose().dot(
                    M1.dot(
                        M2.dot(
                            np.kron(W.transpose(),Iq)
                        )+np.kron(W,Ip).dot(
                            per_matrix(p,q)
                        )
                    )
                ).dot(
                    np.kron(X[j].transpose(),Iq)
                )

                delta2 = M4.dot(
                    np.kron(
                        (W.dot(
                            W.transpose())),(M2.dot(
                                np.kron(W.transpose(),Iq)
                            )+np.kron(W,Ip).dot(
                                    per_matrix(p,q)
                            )
                        )
                    )
                    +np.kron(Ip,np.kron(W.dot(W.transpose()),Ip)).dot(
                        np.kron(per_matrix(p,p),Ip).transpose().dot(
                            np.kron(
                                Ip,M2.dot(
                                    np.kron(W.transpose(),Iq)
                                )+np.kron(W,Ip).dot(
                                    per_matrix(p,q)
                                )
                            ).dot(
                                np.kron(per_matrix(p,p),Iq)
                            )
                        )
                    )
                ).dot(M7)

                delta3 = -2*(
                    M4.dot(
                        M5.dot(
                            np.kron(Ip,M2).dot(
                                M8.dot(
                                    np.kron(vec_by_col(W.transpose()),Iq)
                                )
                            )
                        )
                    )
                    +np.kron(vec_by_col(W).transpose(),Ip).dot(
                        M9.dot(
                            M7
                        )
                    )
                )

                #Updates the weight matrix
            #W = W - l_rate*(delta1)
            #W = W - l_rate*(delta1+delta2+delta3)
            W = W - l_rate*(calibration_epsilon*delta1 + calibration_etha*(delta2 + delta3))
        if(np.math.isnan(c_errors)):
            print("tambien nan")
            break
        else:
            errors2.append(np.trace((W.transpose().dot(W)-Iq).transpose().dot(W.transpose().dot(W)-Iq)))
            errors.append(c_errors/n)
#==============================================================================    


#==============================================================================    
#Finishes program
#==============================================================================
os.system('play --no-show-progress --null --channels 1 synth %s sine %f' % (duration, freq))
#==============================================================================    


#==============================================================================    
#Plotting results
#==============================================================================
plt.plot([np.mean(errors2[i-1:i]) for i in range(len(errors2))], label = 'MSE')
plt.xlabel("epochs")
plt.ylabel("MSE")
plt.legend()
plt.show()
print(W)
print("last error")
print(len(errors))
print(errors.pop())
print((errors2.pop()))
print((errors2.pop())+errors.pop())
print("-")
print(W.transpose().dot(W))
print("")
print(W.transpose().dot(W)-Iq)
#==============================================================================            