import tensorflow as tf
import openpyxl as oxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

import os
duration = 1  # second
freq = 440  # Hz

color = ['b','r','g','c','m','y','k']
type_line = ['', '--', ':', '-.']

l_rate_list = [1E-4, 5E-5, 1E-5]
weightNorm1 = 10
weightOrt = 10000
epsilon = 0.000001
data_set_ = [1,2,3]

for data_set in data_set_:
  #==============================================================================    
  #Data
  #==============================================================================
  if(data_set == 1):
    f_name = 'GeneratedDataThesis'
    wb = oxl.Workbook()
    wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/GeneratedDataThesis.xlsx')
    ws = wb.active
    df = pd.DataFrame(ws.values)
    df = df.drop(df.columns[[0]], axis=1)
    df = df.drop(df.index[0])
    df_X = df.drop(df.columns[[3, 4]], axis=1)
    df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

  elif(data_set == 2):
    f_name = 'bad-drivers_'
    wb = oxl.Workbook()
    wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/bad-drivers_.xlsx')
    ws = wb.active
    df = pd.DataFrame(ws.values)
    df = df.drop(df.columns[[0]], axis=1)
    df = df.drop(df.index[0])
    df_X = df.drop(df.columns[[3, 4, 5]], axis=1)
    df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

  elif(data_set == 3):
    f_name = 'FB-HP'
    wb = oxl.Workbook()
    wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/FB-HP.xlsx')
    ws = wb.active
    df = pd.DataFrame(ws.values)
    df = df.drop(df.columns[[0]], axis=1)
    df = df.drop(df.index[0])
    df_X = df.drop(df.columns[[3, 4]], axis=1)
    df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)


  print('==============================================')
  print(f_name)

  #Converts the data frames to matrices
  X_matrix = np.asmatrix(df_X)
  Y_matrix = np.asmatrix(df_Y)

  #==============================================================================    
  #First variables
  #==============================================================================
  #Sizes
  p = X_matrix.shape[1] #p
  q = Y_matrix.shape[1] #q
  N = X_matrix.shape[0] #n

  #Placeholders for tensors that will be always fed
  X = tf.placeholder(tf.float32, [None,p]) 
  Y = tf.placeholder(tf.float32, [None,q])
  dict = {X: X_matrix, Y: Y_matrix}
  
  #==============================================================================    
  #Functions
  #==============================================================================
  #Lists of variables for the q vectors w
  def CreateListsVarsWs(p,q):
    with tf.name_scope('weights'):
      varsWs_ = [tf.Variable(tf.truncated_normal([p,1], stddev=0.001))]
      for _ in range(1,q):
        varsWs_.append(tf.Variable(tf.truncated_normal([p,1], stddev=0.001)))
      return varsWs_

  #Lists of placeholders for the first q-1 vectors w to be found
  def CreatephWs(p,q):
    phWs_  = [tf.placeholder(tf.float32, [p,None])]
    for _ in range(1,q):
      phWs_.append(tf.placeholder(tf.float32, [p,None]))              
    return phWs_

  #Models and loss functions
  def CreateModels(varsWs_):
    #Gets the variable of the model (w_i)
    W = varsWs_[0]
    #RNA's model: X(W)(WtXY) = Y
    models_ = [tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W),tf.transpose(W)),tf.transpose(X)),Y)] 
    #Sum of the squared errors
    losses_model_ = [tf.reduce_sum(tf.square(Y - models_[0]))]  
    #Norm of the vector equal to one, with a weight to increase its importance
    losses_norm_ = [weightNorm1*tf.square(tf.sqrt(tf.reduce_sum(tf.multiply(W,W))) - 1)]
    #Initializes for the others q-1 Y-variables
    for i in range(1,q):
      W = varsWs_[i]
      models_.append(tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W),tf.transpose(W)),tf.transpose(X)),Y))
      losses_model_.append(tf.reduce_sum(tf.square(Y - models_[i])))
      losses_norm_.append(weightNorm1*tf.square(tf.sqrt(tf.reduce_sum(tf.multiply(W,W))) - 1))
    return losses_model_, losses_norm_

  #Losses from orthonormality with every combination of the w_i
  def OrthonormalLoss(varsWs_,phWs_):
    losses_orth_ = [weightOrt*tf.square(tf.reduce_sum(tf.multiply(varsWs_[1],phWs_[0])))]
    for i in range(2,q):
      for j in range(0,i):
        losses_orth_.append(weightOrt*tf.square(tf.reduce_sum(tf.multiply(varsWs_[i],phWs_[j]))))
    return losses_orth_

  #Final losses (squared sum of the error, normality of the vector, orthonormality with the previouses)
  def ListTotalLost(losses_model,losses_norm,losses_orth):
    loss_total_ = [losses_model[0] + losses_norm[0]]
    for i in range(1,q):
      k = int((i)*(i-1)/2)
      orthonormality = 0
      for j in range(k, k + i):
        orthonormality = orthonormality + losses_orth[j]
      loss_total_.append(losses_model[i] + losses_norm[i] + orthonormality)
    return loss_total_

  #Optimizers and lists of errors
  def Optimizer(total_losses, l_rate):
    globalStep_ = [tf.Variable(0, trainable=False)]
    l_rate_ = [tf.train.exponential_decay(l_rate, globalStep_[0], 1000000, 0.96, staircase=True)]
    optimizer_ = [tf.train.GradientDescentOptimizer(l_rate_[0])]
    train_ = [optimizer_[0].minimize(total_losses[0])]
    errors_ = [[]]
    for i in range(1,q):
      globalStep_.append(tf.Variable(0, trainable=False))
      l_rate_.append(tf.train.exponential_decay(l_rate, globalStep_[i], 1000000, 0.96, staircase=True))
      optimizer_.append(tf.train.GradientDescentOptimizer(l_rate_[i]))
      train_.append(optimizer_[i].minimize(total_losses[i]))
      errors_.append([])
    return train_, errors_

  #Constructs a hyperparameter string dor each one (example: "lr_1E-3")
  def Make_hparam_string(l_rate_):
    return "lr_%s"%l_rate_

  def Neural_net(learning_rate_, name_, run_):
    print('==============================================')
    print(name_)
    print('==============================================')
    #==============================================================================    
    #Hyperparameters
    #==============================================================================
    #initial variables 
    l_rate = learning_rate_
    hparam_str = name_

    #==============================================================================    
    #Initialization of lists, models and loss functions
    #==============================================================================
    varsWs = CreateListsVarsWs(p,q)
    phWs = CreatephWs(p,q)
    losses_model, losses_norm = CreateModels(varsWs)
    losses_orth = OrthonormalLoss(varsWs,phWs)
    total_losses = ListTotalLost(losses_model, losses_norm, losses_orth)
    trains, errors = Optimizer(total_losses, l_rate)

    #==============================================================================    
    #Training
    #==============================================================================
    #initialise the global variables in order to run the session
    init = tf.global_variables_initializer()
    sess = tf.Session()
    sess.run(init)

    #Iterations over the w_i
    for i in range(0,q):
      w_old = np.random.normal(0,0.1,p).reshape(p,1)
      w = w_old
      aux = True

      #Training of w_i
      while((np.sqrt(np.sum((w_old - w).transpose().dot(w_old - w)))) > epsilon  or aux):
        w_old = w
        aux = False
        _, loss_e, w = sess.run([trains[i], total_losses[i], varsWs[i]], dict)
        errors[i].append(loss_e/N)
        
      os.system('play --no-show-progress --null --channels 1 synth %s sine %f' % (duration, freq))
      
      #Saves the value of the w_i found
      if(i == 0):
        Wfinal = [w]
      else:
        Wfinal.append(w)

      #Appends the w_i found its placeholder in the dictionary 
      if((i+1) < q):
        dict[(phWs[i])] = w

    sess.close()

    #==============================================================================    
    #Results
    #==============================================================================
    final_W = np.zeros((p,q)).reshape(p,q)
    for j in range(0,q):
      for i in range(0,p):
        final_W[i,j] = np.asarray((Wfinal[j])[i])

    print("")
    print("W:")
    print(final_W)
    print("IR: %s"%np.trace(final_W.transpose().dot(X_matrix.transpose()).dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W)))
    for i in range(q):
      print('cor²(w%s'%(i+1)+'): %s'%np.trace(final_W[:,i].transpose().dot(X_matrix.transpose()).dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W[:,i])))

    os.system('play --no-show-progress --null --channels 1 synth %s sine %f' % (duration, freq))

    final_W = np.asmatrix(final_W)
    for n in range(q):
      print("")
      print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W[:,n])/final_W[:,n])
    print('==============================================')
    print('')

    #==============================================================================    
    #Graphs
    #==============================================================================
    ax = plt.subplot()
    sizes = [0]
    size = 0
    for i in range(q):
      size = size + len(errors[i])
      sizes.append(size)
    t = []
    for i in range(q):
      t.append(np.arange(sizes[i], sizes[i+1], 1))
    names = []
    for i in range(q):
      names.append('w%s'%(i+1))
    for i in range(q):
        ax.plot(t[i], errors[i], color[i]+type_line[i], label = names[i])
    ax.plot([500], [min(errors[0])], 'w^', label = hparam_str)
    ax.grid(color='k', linestyle='-', linewidth=0.1)
    ax.legend(loc = 'lower right', frameon = False)
    plt.rc('figure', titlesize=12)
    plt.title("Funciones de costo", fontsize = 15, fontweight = 'bold')
    plt.xlabel('Épocas', fontsize = 12, fontweight = 'bold')
    plt.ylabel('Pérdida', fontsize = 12, fontweight ='bold')
    #plt.savefig('/home/ferjyanez/Dropbox/Tesis AR-RNA Sin tutores/Tomo/Data/Graphs/'+f_name+'_'+hparam_str+'.png')
    #plt.show()
    plt.close()

  #End of the Nerual_net function


  for l_rate in l_rate_list:
    hparam_str = Make_hparam_string(l_rate)
    Neural_net(l_rate, hparam_str, l_rate_list.index(l_rate))
