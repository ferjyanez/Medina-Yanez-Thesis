import tensorflow as tf
import openpyxl as oxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter


import os
duration = 1  # second
freq = 440  # Hz

#initial variables 
l_rate_0 = 0.000001

epochs1 = 10000
epochs2 = 70000
epochs3 = 200000

opcion = 9

class NodeTensor:
  def __init__(self, Tensor):
    self.tensor = Tensor

  def setNext(self, Tensor):
    self.next = Tensor
  def getNext(self):
    return self.next
  def setTensor(self, Tensor):
    self.tensor = Tensor
  def getTensor(self):
    return self.tensor

class ListTensor:
  def __init__(self,TensorHead):
    self.head = NodeTensor(TensorHead)
  
  def count(self):
    Var = self.head
    N = 0
    while(Var is not None):
      N = N + 1
      Var = Var.getNext()
    return N
  def getPosition(self,n):
    Var=self.head
    while(n>0):
      n=n-1
      Var=Var.getNext()
    return Var.tensor
  def insert(self,Tensor):
    Node=NodeTensor(Tensor)
    Var=self.head
    while(Var.getNext is not None):
      Var=Var.getNext()
    Var.setNext(Node)



wb = oxl.Workbook()
if(opcion == 1):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/MyData.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:200]
  
  df_X = df.drop(df.columns[[4, 5]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 3]], axis=1)

elif(opcion == 2):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/FB-HP.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:99]

  df_X = df.drop(df.columns[[3, 4]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

elif(opcion == 3):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/hate_crimes.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:50]

  df_X = df.drop(df.columns[[9, 10]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 3, 4, 5, 6, 7, 8]], axis=1)

elif(opcion == 4):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/FB-HP.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:99]

  df_X = df.drop(df.columns[[2, 3, 4]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 4]], axis=1)

elif(opcion == 5):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/Ejemplito.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[2]], axis=1)
  df_Y = df.drop(df.columns[[0, 1]], axis=1)

elif(opcion == 6):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/Ejemplito2.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[2, 3]], axis=1)
  df_Y = df.drop(df.columns[[0, 1]], axis=1)

elif(opcion == 7):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/ejemplito4_.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[3]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)

elif(opcion == 8):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/ejemplito5_.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:4]

  df_X = df.drop(df.columns[[4]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2, 3]], axis=1)

elif(opcion == 9):
  wb = oxl.load_workbook('/home/ferjyanez/Documents/Tesis/Medina-Yanez-Thesis/data/Weather_.xlsx')
  ws = wb.active
  df = pd.DataFrame(ws.values)

  df = df.drop(df.columns[[0]], axis=1)
  df = df.drop(df.index[0])
  df = df[:365]

  df_X = df.drop(df.columns[[3, 4, 5]], axis=1)
  df_Y = df.drop(df.columns[[0, 1, 2]], axis=1)


X_matrix = np.asmatrix(df_X)
Y_matrix = np.asmatrix(df_Y)

#Sizes
p = X_matrix.shape[1] #p
q = Y_matrix.shape[1] #q
n = X_matrix.shape[0] #n

X = tf.placeholder(tf.float32, [None,p]) #Inserts a placeholder for a tensor that will be always fed.
Y = tf.placeholder(tf.float32, [None,q]) #Inserts a placeholder for a tensor that will be always fed.
listVarW, listPHW = createList(p)

""" W_1_ph = tf.placeholder(tf.float32, [p,None]) #gets a tensor with the values
W_2_ph = tf.placeholder(tf.float32, [p,None]) #gets a tensor with the values

W_1 = tf.Variable(tf.truncated_normal([p,1], stddev=0.01)) #gets a tensor with the values
W_2 = tf.Variable(tf.truncated_normal([p,1], stddev=0.01)) #gets a tensor with the values
W_3 = tf.Variable(tf.truncated_normal([p,1], stddev=0.01)) #gets a tensor with the values
#W = tf.Variable(tf.truncated_normal([p,3], stddev=0.01)) #gets a tensor with the values """

def createList(p):
  n = 1
  listVarW_ = ListTensor(tf.Variable(tf.truncated_normal([p,1], stddev=0.01)))
  listPHW_ = ListTensor(tf.placeholder(tf.float32, [p,None]))
 
  while(n < p):
    listVarW_.insert(tf.Variable(tf.truncated_normal([p,1], stddev=0.01)))
    if(n != (p-1)):
      listPHW_.insert(tf.placeholder(tf.float32, [p,None]))
    n = n + 1
  return listVarW_, listPHW_


#Cost functions
def createModels(listVarW):
  n = 1
  W = listVarW.getPosition(0)
  listModels_ = ListTensor(tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W),tf.transpose(W)),tf.transpose(X)),Y))
  
 
  while(n < p):
    W=W.getNext().getTensor()
    listModels_.insert(tf.Variable(tf.truncated_normal([p,1], stddev=0.01)))
    if(n != (p-1)):
      listPHW_.insert(tf.placeholder(tf.float32, [p,None]))
    n = n + 1
  return listVarW_, listPHW_



model1 = tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W_1),tf.transpose(W_1)),tf.transpose(X)),Y) #Y estimada
cost1 = tf.square(Y - model1) #subtracts element-wise the vector and the elevates them to the 2nd power
loss1 = tf.reduce_sum(cost1)
loss1_n = tf.square(tf.sqrt(tf.reduce_sum(tf.multiply(W_1,W_1))) - 1)


model2 = tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W_2),tf.transpose(W_2)),tf.transpose(X)),Y) #Y estimada
cost2 = tf.square(Y - model2) #subtracts element-wise the vector and the elevates them to the 2nd power
loss2 = tf.reduce_sum(cost2)
loss2_n = tf.square(tf.sqrt(tf.reduce_sum(tf.multiply(W_2,W_2))) - 1)
#loss2 = 0

model6 = tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,W_3),tf.transpose(W_3)),tf.transpose(X)),Y) #Y estimada
cost6 = tf.square(Y - model6) #subtracts element-wise the vector and the elevates them to the 2nd power
loss6 = tf.reduce_sum(cost6)
loss6_n = tf.square(tf.sqrt(tf.reduce_sum(tf.multiply(W_3,W_3))) - 1)
#loss6 = 0

loss3 = 10000*tf.square(tf.reduce_sum(tf.multiply(W_1_ph,W_2)))
#loss3 = 10000*tf.square(tf.reduce_sum(tf.multiply(tf.divide(W_1_ph, tf.sqrt(tf.reduce_sum(tf.multiply(W_1_ph,W_1_ph)))),tf.divide(W_2, tf.sqrt(tf.reduce_sum(tf.multiply(W_2,W_2)))))))
#loss3 = 20*(tf.matmul(tf.transpose(W_1), W_2))[0,0]
loss4 = 10000*tf.square(tf.reduce_sum(tf.multiply(W_1_ph,W_3)))
#loss4 = 100000*tf.square(tf.reduce_sum(tf.multiply(tf.divide(W_1_ph, tf.sqrt(tf.reduce_sum(tf.multiply(W_1_ph,W_1_ph)))),tf.divide(W_3, tf.sqrt(tf.reduce_sum(tf.multiply(W_3,W_3)))))))
#loss4 = 40*(tf.matmul(tf.transpose(W_1), W_3))[0,0]
loss5 = 10000*tf.square(tf.reduce_sum(tf.multiply(W_2_ph,W_3)))
#loss5 = 100000*tf.square(tf.reduce_sum(tf.multiply(tf.divide(W_2_ph, tf.sqrt(tf.reduce_sum(tf.multiply(W_2_ph,W_2_ph)))),tf.divide(W_3, tf.sqrt(tf.reduce_sum(tf.multiply(W_3,W_3)))))))
#loss5 = 100*(tf.matmul(tf.transpose(W_2), W_3))[0,0]

loss2_1 = 0 

def f_W(X1, X2, X3, X4):
  X1[:,0].assign(X2)
  X1[:,1].assign(X3)
  X1[:,2].assign(X4)
  return X1

#W = f_W(W, W_1, W_2)

#model3 = tf.matmul(tf.matmul(tf.matmul(tf.matmul(X,f_W(W, W_1, W_2, W_3)),tf.transpose(f_W(W, W_1, W_2, W_3))),tf.transpose(X)),Y)
#cost3_0 = tf.square(Y - model3) 
#cost3_1 = 100*tf.reduce_sum(cost3_0)
#cost3_1 = 0

#cost3 = 0


#loss2 = tf.trace(tf.matmul(tf.transpose(tf.subtract(tf.matmul(tf.transpose(W_1),W_2),tf.eye(1))),tf.subtract(tf.matmul(tf.transpose(W_1),W_2),tf.eye(1))))
#loss2 = tf.reduce_sum(tf.square(tf.matmul(tf.transpose(W_1),W_2)-1))                  



#cost4 = tf.trace(tf.matmul(tf.transpose(tf.subtract(tf.matmul(tf.transpose(f_W(W, W_1, W_2)),f_W(W, W_1, W_2)),tf.eye(q))),tf.subtract(tf.matmul(tf.transpose(f_W(W, W_1, W_2)),f_W(W, W_1, W_2)),tf.eye(q))))
cost4 = 0

total_loss1 = loss1 + loss1_n
total_loss2 = loss2 + loss3 + loss2_n
total_loss3 = loss6 + loss4 + loss5 + loss6_n

#loss = tf.add(tf.add(tf.add(tf.add(tf.add(loss1, loss2), loss3),cost3_1), loss4), loss5)

#Optimizer
#global_step = tf.Variable(0, trainable=False)
#l_rate = tf.train.exponential_decay(l_rate_0, global_step, 10000, 0.96, staircase=True)
optimizer = tf.train.GradientDescentOptimizer(l_rate_0)
train1 = optimizer.minimize(total_loss1)
train2 = optimizer.minimize(total_loss2)
train3 = optimizer.minimize(total_loss3)


errors_e1 = []
errors_e2 = []
errors_e3 = []

#training
#initialise the global variables in order to run the session
init = tf.global_variables_initializer()
sess = tf.Session()
sess.run(init)
for i in range(epochs1):
  _, loss_e1, w1 = sess.run([train1, total_loss1, W_1], {X: X_matrix, Y: Y_matrix})
  errors_e1.append(loss_e1/n)
for i in range(epochs2):
  _, loss_e2, w2 = sess.run([train2, total_loss2, W_2], {X: X_matrix, Y: Y_matrix, W_1_ph: w1})
  errors_e2.append(loss_e2/n)
for i in range(epochs3):
  _, loss_e3, w3 = sess.run([train3, total_loss3, W_3], {X: X_matrix, Y: Y_matrix, W_1_ph: w1, W_2_ph: w2})
  errors_e3.append(loss_e3/n)



#Get results

#_, final_W1, final_W2, final_W3, loss_eu, loss_3, loss_4, loss_5  = sess.run([train, W_1, W_2, W_3, loss, loss3, loss4, loss5], {X:X_matrix, Y:Y_matrix})

print("W3t W2:")
print(w3.transpose().dot(w2))

#final_W1 = final_W1/(np.sqrt(final_W1.transpose().dot(final_W1)))
#final_W2 = final_W2/(np.sqrt(final_W2.transpose().dot(final_W2)))
#final_W3 = final_W3/(np.sqrt(final_W3.transpose().dot(final_W3)))

final_W = np.zeros((p,3)).reshape(p,3)
final_W[0,0] = np.asarray(w1[0])
final_W[1,0] = np.asarray(w1[1])
final_W[2,0] = np.asarray(w1[2])
final_W[0,1] = np.asarray(w2[0])
final_W[1,1] = np.asarray(w2[1])
final_W[2,1] = np.asarray(w2[2])
final_W[0,2] = np.asarray(w3[0])
final_W[1,2] = np.asarray(w3[1])
final_W[2,2] = np.asarray(w3[2])

print("")
print("W:")
print(final_W)
print("")
print("W1:")
print(w1)
print("")
print("W2:")
print(w2)
print("")
print("W3:")
print(w3)
print("")
print("loss: %s"%((loss_e1+loss_e2+loss_e3)/n))
print("IR: %s"%np.trace(final_W.transpose().dot(X_matrix.transpose()).dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W)))
print("")
print("Wt W:")
print(final_W.transpose().dot(final_W))
print("")
#print("W1t W2:")
#print(final_W1.transpose().dot(final_W2))

sess.close()

os.system('play --no-show-progress --null --channels 1 synth %s sine %f' % (duration, freq))


print("")
#W_AR = [[-0.9209633,0.3896493],[-0.3896493,-0.9209633]]
#W_AR = np.asmatrix(W_AR)
final_W = np.asmatrix(final_W)
#print(final_W[2,0]/W_AR[2,0])
print("")
print("final_W")
#print(W_AR)
print("")
print(final_W)
print("")
#print("")
#print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(W_AR[:,0])/W_AR[:,0])
#print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(W_AR[:,1])/W_AR[:,1])
print("")
print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W[:,0])/final_W[:,0])
print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W[:,1])/final_W[:,1])
print(X_matrix.transpose().dot(Y_matrix).dot(Y_matrix.transpose()).dot(X_matrix).dot(final_W[:,2])/final_W[:,2])

#plt.subplot(1,1,1)
plt.plot(np.asarray(errors_e1+errors_e2+errors_e3))
plt.ylabel("Loss")
plt.show()
#plt.savefig("errors.png")

"""

W_AR = [[0.5544050,0.821733],[-0.1993123,1.2850765],[-0.8080283,0.4934761]]
W_AR = np.asmatrix(W_AR)
print("")
print(W_AR[0:3,0])
print("")
print(np.cross(W_AR[:,0].transpose(), final_W[:,0].transpose()))
print("")
print(W_AR[0,0]/final_W[0,0])
print(W_AR[1,0]/final_W[1,0])
print(W_AR[2,0]/final_W[2,0])




def f(w1, w2, X_matrix, Y_matrix):

  Z = np.zeros((20,20))
  for i in range(20):
    for j in range(20):
      Z[i,j] = (Y_matrix - (np.multiply(np.square(w1[i,j]),X_matrix[:,0])
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,1])).dot(
                  X_matrix[:,0].transpose()
                ).dot(
                  Y_matrix)
              + (np.multiply(np.square(w2[i,j]),X_matrix[:,1]) 
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,0])).dot(
                  X_matrix[:,1].transpose()
                ).dot(
                    Y_matrix)).transpose().dot(
      Y_matrix - (np.multiply(np.square(w1[i,j]),X_matrix[:,0])
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,1])).dot(
                  X_matrix[:,0].transpose()
                ).dot(
                  Y_matrix[:,0])
              + (np.multiply(np.square(w2[i,j]),X_matrix[:,1]) 
              + np.multiply(np.multiply(w1[i,j],w2[i,j]),X_matrix[:,0])).dot(
                  X_matrix[:,1].transpose()
                ).dot(
                    Y_matrix)
              )
  return Z

# Make data.
fig = plt.figure()
ax = fig.gca(projection='3d')

w1 = np.linspace(-1.3, 1.3, 20)
w2 = np.linspace(-1.3, 1.3, 20)
w1, w2 = np.meshgrid(w1, w2)

Z = np.multiply(f(w1, w2, X_matrix, Y_matrix),(1/n))


# Plot the surface.
surf = ax.plot_surface(w1, w2, Z, cmap = cm.coolwarm,linewidth=0, antialiased=False)

# Customize the z axis.
ax.set_zlim(0, 2.5)
ax.zaxis.set_major_locator(LinearLocator(10))
ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))

# Add a color bar which maps values to colors.
fig.colorbar(surf, shrink=0.5, aspect=5)

plt.show()

 """